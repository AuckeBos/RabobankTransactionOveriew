import urllib2
from BeautifulSoup import BeautifulSoup
from xlwt import Workbook
import xlrd
from xlutils.copy import copy
import xlutils.copy
import openpyxl as openpyxl
from subprocess import call
import sys
from data import Data
import os.path
from datetime import date

def getMonth(month):
    months = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
    codes = ['{year:02d}-{month:02d}-01'.format(year=date.today().year, month=i) for i in range(1, 13)]
    if month in months:
        return codes[months.index(month)]
    print "Month not found! please choose from {}".format(months.lower())
    return getMonth(str(raw_input("Choose month: ")))

def createSheet():
    if not "Template" in file.sheetnames:
        sys.exit("Template sheet not found! please create it")
    template = file.get_sheet_by_name("Template")
    newSheet = file.copy_worksheet(template)
    newSheet.title = sheetName
    file.save(fileName)
    print 'Sheet {} added to \'Kosten - inkomsten 2018.xlsx\''.format(sheetName)

def fillSheet(allData, exclude):
    sheet = file.get_sheet_by_name(sheetName)
    positiveTransactions = []
    negativeTransactions = []
    boodschappenAH = []
    pinnen = []
    for row in allData:
        transaction = {}
        date = row.find("span", {"class":"name", "id":"valueDate"})
        acountNumber = row.find("span", {"class":"accountnumber"})
        description = row.find("div", {"class":"description"})
        name = row.find("span", {"class":"name"})
        amount = row.find("td", {"class":"amount "})

        if date != None and acountNumber != None and not acountNumber.text in exclude:
            transaction["name"] = "" if name == None else str(name.text)
            transaction["date"] = str(date.text)
            transaction["description"] = str(description.text)[:30]
            transaction["amount"] = float(str(amount.text).replace(".", "").replace(",", ".").replace("-", ""))
            if "-" in str(amount.text):
                negativeTransactions.append(transaction)
            else:
                positiveTransactions.append(transaction)
            if "ALBERT" in str(transaction["name"]):
                boodschappenAH.append(transaction)
            elif "Geldautomaat" in str(description.text):
                pinnen.append(transaction)

    for i in range (len(negativeTransactions)):
        sheet.cell(row = i+3, column = 1).value = negativeTransactions[i]['date']
        sheet.cell(row = i+3, column = 2).value = negativeTransactions[i]['amount']
        sheet.cell(row = i+3, column = 3).value = negativeTransactions[i]['description']
        sheet.cell(row = i+3, column=4).value = negativeTransactions[i]['name']
    for i in range (len(positiveTransactions)):
        sheet.cell(row = i+3, column = 6).value = positiveTransactions[i]['date']
        sheet.cell(row = i+3, column = 7).value = positiveTransactions[i]['amount']
        sheet.cell(row = i+3, column = 8).value = positiveTransactions[i]['description']
        sheet.cell(row = i+3, column=9).value = positiveTransactions[i]['name']
    for i in range (len(boodschappenAH)):
        sheet.cell(row = i+14, column = 11).value = boodschappenAH[i]['amount']
    for i in range (len(pinnen)):
        sheet.cell(row = i+14, column = 12).value = pinnen[i]['amount']
    file.save(fileName)
    print '{} transactions added to {} in \'Kosten - inkomsten 2017.xlsx\''.format(len(allData), sheetName)

if not os.path.exists('Kosten - inkomsten 2018.xlsx'):
    sys.exit("File Kosten - inkomsten 2018.xlsx not found!")

sheetName = str(raw_input("Choose month: "))
month = getMonth(sheetName.lower())
fileName = "Kosten - inkomsten 2018.xlsx"
file = openpyxl.load_workbook(fileName)

createSheet()
data, exclude = Data().getData(month)
fillSheet(data, exclude)
